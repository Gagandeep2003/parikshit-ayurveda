"""
Shree Parikshit Ayurveda - Clinic & Panchakarma Centre
Full-stack Flask Application
"""

import os
import sqlite3
import hashlib
import secrets
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from datetime import datetime, timedelta
from functools import wraps
from io import BytesIO

from flask import (Flask, render_template, request, redirect, url_for,
                   session, flash, jsonify, send_file)
from apscheduler.schedulers.background import BackgroundScheduler
import openpyxl
import pytz

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', secrets.token_hex(32))

# Configuration
CLINIC_EMAIL = 'bharpur.spayur@gmail.com'
CLINIC_PHONE = '9990978771'
CLINIC_WHATSAPP = '919990978771'
SMTP_EMAIL = os.environ.get('SMTP_EMAIL', CLINIC_EMAIL)
SMTP_PASSWORD = os.environ.get('SMTP_PASSWORD', '')
SMTP_HOST = os.environ.get('SMTP_HOST', 'smtp.gmail.com')
SMTP_PORT = int(os.environ.get('SMTP_PORT', 587))

IST = pytz.timezone('Asia/Kolkata')

# ─── Database Setup ────────────────────────────────────────────────
DB_PATH = os.environ.get('DB_PATH', 'clinic.db')

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db()
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        email TEXT UNIQUE NOT NULL,
        phone TEXT NOT NULL,
        address TEXT DEFAULT '',
        password TEXT NOT NULL,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS bookings (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER,
        name TEXT NOT NULL,
        email TEXT NOT NULL,
        phone TEXT NOT NULL,
        date TEXT NOT NULL,
        time_slot TEXT NOT NULL,
        concern TEXT DEFAULT '',
        status TEXT DEFAULT 'confirmed',
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (user_id) REFERENCES users(id),
        UNIQUE(date, time_slot)
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS contacts (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        email TEXT NOT NULL,
        phone TEXT DEFAULT '',
        message TEXT NOT NULL,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )''')
    conn.commit()
    conn.close()

init_db()

# ─── Helpers ───────────────────────────────────────────────────────
def hash_password(password):
    return hashlib.sha256(password.encode()).hexdigest()

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            flash('Please login to continue.', 'warning')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated

def get_time_slots():
    """Generate 30-minute time slots from 10 AM to 8 PM"""
    slots = []
    hour = 10
    minute = 0
    while hour < 20:
        start = f"{hour:02d}:{minute:02d}"
        end_min = minute + 30
        end_hour = hour
        if end_min >= 60:
            end_min -= 60
            end_hour += 1
        end = f"{end_hour:02d}:{end_min:02d}"
        slots.append(f"{start} - {end}")
        minute += 30
        if minute >= 60:
            minute -= 60
            hour += 1
    return slots

TIME_SLOTS = get_time_slots()

def get_booked_slots(date_str):
    conn = get_db()
    rows = conn.execute(
        'SELECT time_slot FROM bookings WHERE date = ? AND status = ?',
        (date_str, 'confirmed')
    ).fetchall()
    conn.close()
    return [r['time_slot'] for r in rows]

# ─── Email with Excel ─────────────────────────────────────────────
def generate_users_excel():
    """Generate Excel file of all registered users"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Registered Users"

    # Header styling
    from openpyxl.styles import Font, PatternFill, Alignment
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="6B2D2D", end_color="6B2D2D", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    headers = ['S.No', 'Name', 'Email', 'Phone', 'Address', 'Registered On']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    conn = get_db()
    users = conn.execute('SELECT * FROM users ORDER BY created_at DESC').fetchall()
    conn.close()

    for idx, user in enumerate(users, 1):
        ws.cell(row=idx+1, column=1, value=idx)
        ws.cell(row=idx+1, column=2, value=user['name'])
        ws.cell(row=idx+1, column=3, value=user['email'])
        ws.cell(row=idx+1, column=4, value=user['phone'])
        ws.cell(row=idx+1, column=5, value=user['address'])
        ws.cell(row=idx+1, column=6, value=user['created_at'])

    # Auto-width columns
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_len + 4

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def send_daily_email():
    """Send daily Excel report at 9 PM IST"""
    if not SMTP_PASSWORD:
        print("[WARN] SMTP_PASSWORD not set. Skipping email.")
        return

    now_ist = datetime.now(IST)
    date_str = now_ist.strftime('%Y-%m-%d')

    try:
        excel_buffer = generate_users_excel()

        msg = MIMEMultipart()
        msg['From'] = SMTP_EMAIL
        msg['To'] = CLINIC_EMAIL
        msg['Subject'] = f'Shree Parikshit Ayurveda - Registered Users Report ({date_str})'

        body = f"""
        <html>
        <body style="font-family: Arial; color: #333;">
            <h2 style="color: #6B2D2D;">🌿 Shree Parikshit Ayurveda</h2>
            <p>Dear Team,</p>
            <p>Please find attached the daily report of all registered users on the website as of <strong>{date_str}</strong>.</p>
            <br>
            <p style="color: #888;">This is an automated email sent at 9:00 PM IST daily.</p>
        </body>
        </html>
        """
        msg.attach(MIMEText(body, 'html'))

        part = MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        part.set_payload(excel_buffer.read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', f'attachment; filename="registered_users_{date_str}.xlsx"')
        msg.attach(part)

        server = smtplib.SMTP(SMTP_HOST, SMTP_PORT)
        server.starttls()
        server.login(SMTP_EMAIL, SMTP_PASSWORD)
        server.send_message(msg)
        server.quit()
        print(f"[OK] Daily email sent successfully at {now_ist}")
    except Exception as e:
        print(f"[ERROR] Failed to send email: {e}")

# ─── Scheduler ─────────────────────────────────────────────────────
scheduler = BackgroundScheduler(timezone=IST)
scheduler.add_job(send_daily_email, 'cron', hour=21, minute=0)
scheduler.start()

# ─── Routes ────────────────────────────────────────────────────────

@app.route('/')
def home():
    return render_template('home.html')

@app.route('/services')
def services():
    return render_template('services.html')

@app.route('/blog')
def blog():
    return render_template('blog.html')

@app.route('/contact', methods=['GET', 'POST'])
def contact():
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        email = request.form.get('email', '').strip()
        phone = request.form.get('phone', '').strip()
        message = request.form.get('message', '').strip()

        if not all([name, email, message]):
            flash('Please fill all required fields.', 'error')
            return redirect(url_for('contact'))

        conn = get_db()
        conn.execute('INSERT INTO contacts (name, email, phone, message) VALUES (?, ?, ?, ?)',
                     (name, email, phone, message))
        conn.commit()
        conn.close()
        flash('Your message has been sent successfully! We will get back to you soon.', 'success')
        return redirect(url_for('contact'))

    return render_template('contact.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form.get('email', '').strip()
        password = request.form.get('password', '')

        conn = get_db()
        user = conn.execute('SELECT * FROM users WHERE email = ?', (email,)).fetchone()
        conn.close()

        if user and user['password'] == hash_password(password):
            session['user_id'] = user['id']
            session['user_name'] = user['name']
            session['user_email'] = user['email']
            flash(f'Welcome back, {user["name"]}! 🌿', 'success')
            return redirect(url_for('home'))
        else:
            flash('Invalid email or password.', 'error')

    return render_template('login.html')

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        email = request.form.get('email', '').strip()
        phone = request.form.get('phone', '').strip()
        address = request.form.get('address', '').strip()
        password = request.form.get('password', '')
        confirm = request.form.get('confirm_password', '')

        if not all([name, email, phone, password]):
            flash('Please fill all required fields.', 'error')
            return redirect(url_for('register'))

        if password != confirm:
            flash('Passwords do not match.', 'error')
            return redirect(url_for('register'))

        conn = get_db()
        existing = conn.execute('SELECT id FROM users WHERE email = ?', (email,)).fetchone()
        if existing:
            conn.close()
            flash('Email already registered. Please login.', 'error')
            return redirect(url_for('login'))

        conn.execute(
            'INSERT INTO users (name, email, phone, address, password) VALUES (?, ?, ?, ?, ?)',
            (name, email, phone, address, hash_password(password))
        )
        conn.commit()
        conn.close()
        flash('Registration successful! Please login.', 'success')
        return redirect(url_for('login'))

    return render_template('register.html')

@app.route('/logout')
def logout():
    session.clear()
    flash('You have been logged out.', 'info')
    return redirect(url_for('home'))

@app.route('/book', methods=['GET', 'POST'])
@login_required
def book():
    if request.method == 'POST':
        date = request.form.get('date', '')
        time_slot = request.form.get('time_slot', '')
        concern = request.form.get('concern', '').strip()

        if not date or not time_slot:
            flash('Please select a date and time slot.', 'error')
            return redirect(url_for('book'))

        # Check if Monday
        selected_date = datetime.strptime(date, '%Y-%m-%d')
        if selected_date.weekday() == 0:
            flash('Sorry, the clinic is closed on Mondays.', 'error')
            return redirect(url_for('book'))

        # Check past date
        today = datetime.now(IST).date()
        if selected_date.date() < today:
            flash('Cannot book for a past date.', 'error')
            return redirect(url_for('book'))

        conn = get_db()
        user = conn.execute('SELECT * FROM users WHERE id = ?', (session['user_id'],)).fetchone()

        # Check if slot already taken
        existing = conn.execute(
            'SELECT id FROM bookings WHERE date = ? AND time_slot = ? AND status = ?',
            (date, time_slot, 'confirmed')
        ).fetchone()

        if existing:
            conn.close()
            flash('This time slot is already booked. Please choose another.', 'error')
            return redirect(url_for('book'))

        conn.execute(
            'INSERT INTO bookings (user_id, name, email, phone, date, time_slot, concern) VALUES (?, ?, ?, ?, ?, ?, ?)',
            (session['user_id'], user['name'], user['email'], user['phone'], date, time_slot, concern)
        )
        conn.commit()
        conn.close()
        flash('Consultation booked successfully! 🌿 We look forward to seeing you.', 'success')
        return redirect(url_for('my_bookings'))

    return render_template('book.html', time_slots=TIME_SLOTS)

@app.route('/api/available-slots')
@login_required
def available_slots():
    date = request.args.get('date', '')
    if not date:
        return jsonify({'error': 'Date required'}), 400

    selected_date = datetime.strptime(date, '%Y-%m-%d')
    if selected_date.weekday() == 0:
        return jsonify({'closed': True, 'message': 'Clinic is closed on Mondays'})

    booked = get_booked_slots(date)
    available = [s for s in TIME_SLOTS if s not in booked]
    return jsonify({'closed': False, 'available': available, 'booked': booked})

@app.route('/my-bookings')
@login_required
def my_bookings():
    conn = get_db()
    bookings = conn.execute(
        'SELECT * FROM bookings WHERE user_id = ? ORDER BY date DESC, time_slot ASC',
        (session['user_id'],)
    ).fetchall()
    conn.close()
    return render_template('my_bookings.html', bookings=bookings)

@app.route('/cancel-booking/<int:booking_id>', methods=['POST'])
@login_required
def cancel_booking(booking_id):
    conn = get_db()
    conn.execute(
        'UPDATE bookings SET status = ? WHERE id = ? AND user_id = ?',
        ('cancelled', booking_id, session['user_id'])
    )
    conn.commit()
    conn.close()
    flash('Booking cancelled successfully.', 'info')
    return redirect(url_for('my_bookings'))

# ─── Error Handlers ────────────────────────────────────────────────
@app.errorhandler(404)
def page_not_found(e):
    return render_template('404.html'), 404

# ─── Run ───────────────────────────────────────────────────────────
if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
